import json
import io
import sys
import logging
import urllib.request
import openpyxl
import json

from django.shortcuts import render
from django.http import HttpResponse
from openpyxl.styles import Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf8') #改变标准输出的默认编码
logger = logging.getLogger("django")
from django.shortcuts import render

def index(request):
    return render(request, "index.html")

def search_htm(request, fl_name):
    return render(request, fl_name + '.html')

def getModels(request):
    def Table2Json():
        try:
            fileName = request.GET.get('fileName')
            searchDate = request.GET.get('searchDate')
            dataSheetName = request.GET.get('dataSheetName')
            responseData = handelExcel(fileName, searchDate, dataSheetName)
            responseData['status'] = 1
            return HttpResponse(content= json.dumps(responseData),
            content_type='application/json;charset = utf-8',
                                        status='200',
                                        reason='success',
                                        charset='utf-8')
            #'文件处理完毕'
        except Exception as e:
            return HttpResponse(content= '请校验参数, C:/Users/lenovo/Desktop/excelPy/handelExlce/handelExlce/' + str(request.GET.get('fileName')+'\n或者该文件是否被占用。'), 
            content_type='application/json;charset = utf-8',
                                        status='400',
                                        reason='error',
                                        charset='utf-8')
    response = Table2Json()
    # print("转换成json的数据", resp)
    return response

def handelExcel(fileName, searchDate, dataSheetName):
    # try:
        # excel处理
        # 需要处理的excel文件地址
        # C:/Users/lenovo/Desktop/excelPy/handelExlce/handelExlce/
        file_path = 'C:/Users/lenovo/Desktop/excelPy/'+ str(fileName)
        # '红包发放汇总表从6-8月(2)(1).xlsx'
        # 加载该文件
        workbook = openpyxl.load_workbook(file_path)
        # searchDate = '20200828'
        # dataSheetName = '8月明细'
        dataSheet = workbook[dataSheetName]
        dataSheetRb = dataSheet['B']

        # 如果操作表不存在则添加
        sheetNames = workbook.sheetnames
        activeSheetName = '数据整合表'
        if sheetNames.count(activeSheetName) == 0:
            workbook.create_sheet(activeSheetName)
        else:
            workbook.remove(workbook[activeSheetName])
            workbook.create_sheet(activeSheetName)

        activeSheet = workbook[activeSheetName]
        #查找人员数据
        peoples = {}
        names = []
        value = []

        for index,item in enumerate(dataSheetRb):
            if str(item.value) == searchDate:
                clous = 'I' + str(index + 1)
                name = str(dataSheet[clous].value)
                if name in peoples:
                    peoples[name] = peoples[name] + 1
                else:
                    peoples[name] = 1

        activeSheet['A1'] = '姓名'
        activeSheet['B1'] = '数量'
        activeSheet['C1'] = '金额'
        peopleIndex = 2
        for item,index in peoples.items():

            activeSheet['A' + str(peopleIndex)].value = item
            activeSheet['B' + str(peopleIndex)].value = index
            activeSheet['C' + str(peopleIndex)].value = index*50
            names.append(item)
            value.append(index*50)
            peopleIndex = peopleIndex + 1

        logger.info('file_path')
        dataList = {}
        dataList['name'] = names
        dataList['value'] = value
        workbook.save(file_path)
        logger.info('file_path')
        return dataList
    # except Exception as e:
    #     raise Exception
        # return '文件处理出错，请校验参数是否正确'